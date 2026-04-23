import os
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Configuration
JIRA_URL = os.getenv('JIRA_URL', 'https://your-jira.atlassian.net')
JIRA_USERNAME = os.getenv('JIRA_USERNAME', 'your-username')
JIRA_API_TOKEN = os.getenv('JIRA_API_TOKEN', 'your-api-token')
JIRA_PROJECT = os.getenv('JIRA_PROJECT', 'YOURPROJECT')
JQL_QUERY = os.getenv('JQL_QUERY', f'project = {JIRA_PROJECT} AND issuetype = Bug AND labels = accessibility')

# Custom field IDs (update based on your JIRA instance)
PAGE_FIELD = 'customfield_10001'  # Replace with actual custom field ID for page/URL

def fetch_jira_issues():
    url = f"{JIRA_URL}/rest/api/2/search"
    headers = {
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }
    auth = (JIRA_USERNAME, JIRA_API_TOKEN)
    payload = {
        'jql': JQL_QUERY,
        'fields': ['key', 'summary', 'description', 'status', 'assignee', 'priority', PAGE_FIELD],
        'maxResults': 1000
    }
    response = requests.post(url, headers=headers, auth=auth, json=payload)
    response.raise_for_status()
    return response.json()['issues']

def process_issues(issues):
    global_defects = []
    page_defects = {}
    
    for issue in issues:
        fields = issue['fields']
        page = fields.get(PAGE_FIELD, {}).get('value', '') if PAGE_FIELD in fields else ''
        defect = {
            'Issue ID': issue['key'],
            'Summary': fields.get('summary', ''),
            'Description': fields.get('description', ''),
            'Status': fields.get('status', {}).get('name', ''),
            'Assignee': fields.get('assignee', {}).get('displayName', '') if fields.get('assignee') else '',
            'Priority': fields.get('priority', {}).get('name', ''),
            'Page': page
        }
        if not page:
            global_defects.append(defect)
        else:
            if page not in page_defects:
                page_defects[page] = []
            page_defects[page].append(defect)
    
    return global_defects, page_defects

def create_excel(global_defects, page_defects, filename='accessibility_report.xlsx'):
    wb = Workbook()
    
    # Global Defects sheet
    ws_global = wb.active
    ws_global.title = 'Global Defects'
    headers = ['Issue ID', 'Summary', 'Description', 'Status', 'Assignee', 'Priority']
    ws_global.append(headers)
    for cell in ws_global[1]:
        cell.font = Font(bold=True)
    
    for defect in global_defects:
        row = [defect[h] for h in headers]
        ws_global.append(row)
    
    # Page-specific sheets
    for page, defects in page_defects.items():
        ws = wb.create_sheet(title=page[:31])  # Excel sheet name limit
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for defect in defects:
            row = [defect[h] for h in headers]
            ws.append(row)
    
    wb.save(filename)
    print(f"Excel report generated: {filename}")

if __name__ == '__main__':
    try:
        issues = fetch_jira_issues()
        global_defects, page_defects = process_issues(issues)
        create_excel(global_defects, page_defects)
    except Exception as e:
        print(f"Error: {e}")